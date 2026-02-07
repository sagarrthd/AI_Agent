from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd


def load_stp_template(path: str) -> Optional[pd.DataFrame]:
    if not path:
        return None
    stp_path = Path(path)
    if not stp_path.exists():
        return None
    return pd.read_excel(stp_path)


def load_stp_template_schema(path: str) -> Dict[str, Any]:
    """Return a light schema (sheet/header/columns) for prompt and writers."""
    if not path or not Path(path).exists():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx in range(1, 26):
                values = [str(c).strip() if c is not None else "" for c in next(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))]
                normalized = [v.lower() for v in values if v]
                if any("test" in v and "id" in v for v in normalized) and (
                    any("step" in v for v in normalized) or any("title" in v for v in normalized)
                ):
                    cols = [v for v in values if v]
                    return {"sheet": sheet_name, "header_row": row_idx, "columns": cols}
    finally:
        wb.close()

    return {}
