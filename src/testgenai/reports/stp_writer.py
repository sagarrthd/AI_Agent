from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from copy import copy

from testgenai.models.testcase import TestCase

def write_stp_output(
    template_path: str,
    output_path: str,
    tests: List[TestCase],
    trace_matrix: Dict[str, List[str]],
    trace_sheet_name: str,
) -> None:
    """
    Writes test cases into an existing Excel template preserving formatting.
    """
    
    # 1. Load the Template
    if not template_path or not Path(template_path).exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    output = Path(output_path)
    # Copy template to output to preserve macros/vba/styles
    wb = openpyxl.load_workbook(template_path)
    
    # 2. Fill Test Plan Sheet
    target_sheet = None
    header_map = {}
    found_header = False
    
    # Smart Scan: Look through ALL sheets for a header row
    print("Scanning sheets for Test Case headers...")
    
    field_keywords = {
        "test_id": ["test id", "id", "case id", "identifier", "test_id"],
        "title": ["title", "summary", "test name", "case name", "test title"],
        "description": ["description", "objective", "purpose", "test scenario"],
        "preconditions": ["preconditions", "pre-requisites", "setup", "pinned", "pin state"],
        "steps": ["steps", "actions", "procedure", "test steps", "step description", "debugger action"],
        "expected": ["expected", "expected result", "expected behavior", "writes expected"],
        "requirements": ["requirement", "req id", "traceability", "ref"],
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Scan first 20 rows of each sheet
        for r_idx in range(1, 21):
            row_values = []
            for c in sheet[r_idx]:
                 val = str(c.value).strip().lower() if c.value else ""
                 row_values.append(val)
            
            # Check for critical keywords to identify this as the Test Plan sheet
            # We look for "id" AND ("step" OR "title" OR "description")
            is_header = False
            if any("test id" in v or "case id" in v for v in row_values):
                if any("step" in v for v in row_values) or any("title" in v for v in row_values):
                    is_header = True
            
            if is_header:
                print(f"✓ Found Test Headers in sheet '{sheet_name}' at row {r_idx}")
                target_sheet = sheet
                # Build Map
                for col_idx, cell_value in enumerate(row_values, 1):
                    for field, keywords in field_keywords.items():
                        if any(k in cell_value for k in keywords):
                            if field not in header_map: # First match wins
                                header_map[field] = col_idx
                found_header = True
                break
        if found_header:
            break
            
    if not target_sheet:
        print("⚠ Could not automatically detect Test Plan sheet. Using active sheet.")
        target_sheet = wb.active
        
    _fill_test_sheet(target_sheet, tests, header_map)
    
    # 3. Fill Traceability Sheet (if requested)
    if trace_sheet_name:
        if trace_sheet_name in wb.sheetnames:
            trace_sheet = wb[trace_sheet_name]
        else:
            trace_sheet = wb.create_sheet(trace_sheet_name)
        _fill_trace_sheet(trace_sheet, trace_matrix)
        
    # 4. Save
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output)
        print(f"✓ Saved strict template output to: {output.resolve()}")
    except Exception as e:
        print(f"❌ FAILED to save file to {output}: {e}")
        raise e

def _fill_test_sheet(sheet, tests: List[TestCase], header_map: Dict[str, int]):
    if not header_map:
        print("⚠ No headers mapped. Appending raw data to end of sheet.")
        # Fallback: just write to columns 1, 2, 3...
        header_map = {
            "test_id": 1, "title": 2, "description": 3, 
            "preconditions": 4, "steps": 5, "expected": 6, "requirements": 7
        }
    
    # Find start row (first empty row after header)
    # We can't just use max_row because template might have formatting down to row 1000
    # Let's find the first row where the 'Test ID' column is empty, starting after the header
    
    # Determine the test_id column to check for emptiness
    check_col = header_map.get("test_id", 1)
    
    # Heuristic: Start from row 1, go down until we find data, then find next empty
    # Actually openpyxl max_row is unreliable with styles.
    # Let's assume we append after the last row that has text in the ID column.
    
    last_data_row = 0
    for row in range(1, 5000): # Scan reasonable limit
        val = sheet.cell(row=row, column=check_col).value
        if val:
            last_data_row = row
            
    current_row = last_data_row + 1
    if current_row < 2: current_row = 2 # Safety
    
    print(f"Writing {len(tests)} tests starting at row {current_row}...")

    for test in tests:
        # Prepare content
        steps_text = "\n".join(f"{i+1}. {s.action}" for i, s in enumerate(test.steps))
        expected_text = "\n".join(f"{i+1}. {s.expected}" for i, s in enumerate(test.steps))
        reqs_text = ", ".join(test.requirements)
        
        # Write to mapped columns
        # Helper to safely write
        def write_cell(field_name, value):
            if field_name in header_map:
                col = header_map[field_name]
                cell = sheet.cell(row=current_row, column=col)
                cell.value = value
                # Optional: Copy style from row above if it exists and has style
                # But only if row above is not header? safe to skip specific styling for now to avoid errors

        write_cell("test_id", test.test_id)
        write_cell("title", test.title)
        write_cell("description", test.title) 
        write_cell("preconditions", test.preconditions)
        write_cell("steps", steps_text)
        write_cell("expected", expected_text)
        write_cell("requirements", reqs_text)
        
        current_row += 1


def _fill_trace_sheet(sheet, trace_matrix: Dict[str, List[str]]):
    # simple append
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value="Requirement ID")
    sheet.cell(row=row, column=2, value="Test Cases")
    
    for req, tests in trace_matrix.items():
        row += 1
        sheet.cell(row=row, column=1, value=req)
        sheet.cell(row=row, column=2, value=", ".join(tests))
