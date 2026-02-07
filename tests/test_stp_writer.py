import tempfile
import unittest
from pathlib import Path

import openpyxl

from testgenai.models.testcase import TestCase, TestStep
from testgenai.reports.stp_writer import write_stp_output


class StpWriterTests(unittest.TestCase):
    def test_write_stp_output_updates_template_rows(self) -> None:
        step = TestStep(step_id="S1", action="Do", expected="Ok", requirement_ids=["REQ-1"])
        test = TestCase(
            test_id="TC-1",
            title="Title",
            preconditions="IGN ON",
            steps=[step],
            requirements=["REQ-1"],
        )

        with tempfile.TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "template.xlsx"
            out_path = Path(tmp) / "out.xlsx"

            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Plan"
            sheet.cell(row=1, column=1, value="Test ID")
            sheet.cell(row=1, column=2, value="Title")
            sheet.cell(row=1, column=3, value="Preconditions")
            sheet.cell(row=1, column=4, value="Steps")
            sheet.cell(row=1, column=5, value="Expected Results")
            sheet.cell(row=1, column=6, value="Requirement IDs")
            sheet.cell(row=2, column=1, value="OLD-TC")
            wb.save(template_path)

            write_stp_output(str(template_path), str(out_path), [test], {"REQ-1": ["TC-1"]}, "Trace")

            out_wb = openpyxl.load_workbook(out_path)
            out_sheet = out_wb["Plan"]
            self.assertEqual(out_sheet.cell(row=2, column=1).value, "TC-1")
            self.assertEqual(out_sheet.cell(row=2, column=2).value, "Title")
            self.assertEqual(out_sheet.cell(row=2, column=4).value, "1. Do")

            trace = out_wb["Trace"]
            self.assertEqual(trace.cell(row=1, column=1).value, "Requirement ID")
            self.assertEqual(trace.cell(row=2, column=1).value, "REQ-1")


if __name__ == "__main__":
    unittest.main()
