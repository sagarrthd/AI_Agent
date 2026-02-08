import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def get_dummy_test_cases():
    return [
        {
            "case_number": 1,
            "id": "TC_HIL_001",
            "title": "Starter lockout and engine start consent validation",
            "description": "Validate starter lockout behavior and engine start consent signal transitions.",
            "requirement_id": "REQ-ENG-001",
            "pre_conditions": "1. Key switch OFF.\n2. CAN channel connected and active.\n3. PCAN Explorer opened.",
            "post_conditions": "NA",
            "expected_result": "System publishes valid starter lockout and consent values.",
            "test_log_path": "",
            "test": "HIL-Open Loop",
            "steps": [
                {
                    "step_num": 1,
                    "description": "Key switch On.",
                    "execution": "Confirm that the CAN channel is connected and active.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 2,
                    "description": "Open PCAN Explorer on your PC and start Trace.",
                    "execution": "Open transmit/receive window and monitor incoming frames.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 3,
                    "description": "Verify PGN ENGSC (F0ED) is available and updating.",
                    "execution": "Verify 10F0ED00 is received.",
                    "actual": "",
                    "expected": "Rx frame valid",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-001",
                    "dcr": "",
                },
                {
                    "step_num": 4,
                    "description": "Send ETC7 message\nCAN ID: 18FE4A03\nData Bytes: 00 00 00 00 00 00 00 00",
                    "execution": "Set cycle time (e.g., 100 ms) and click Send.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 5,
                    "description": "Verify starter lockout status data is valid.",
                    "execution": "Read starter lockout status signal.",
                    "actual": "",
                    "expected": "DATA_VALID",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-002",
                    "dcr": "",
                },
                {
                    "step_num": 6,
                    "description": "Send PropA message\nCAN ID: 18EF0027\nData Bytes: 00 00 00 00 00 00 00 00",
                    "execution": "Send 18EF0027 frame repeatedly.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 7,
                    "description": "Verify starter lockout is 0 and status is data valid.",
                    "execution": "Read lockout value and status bits.",
                    "actual": "",
                    "expected": "0",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-003",
                    "dcr": "",
                },
                {
                    "step_num": 8,
                    "description": "Open PCAN Explorer and verify PGN ENGSC (F0ED) updates.",
                    "execution": "Verify 10F0ED00 is received.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 9,
                    "description": "Verify SPN 7746 byte1 bit position 3(1.3) is 000.",
                    "execution": "Observe SPN bit value in monitor.",
                    "actual": "",
                    "expected": "000 - hex\n000 - dec",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-004",
                    "dcr": "",
                },
                {
                    "step_num": 10,
                    "description": "Send ETC7 message\nCAN ID: 18FE4A03\nData Bytes: 00 10 00 00 00 00 00 00",
                    "execution": "Set cycle time (e.g., 100 ms) and click Send.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 11,
                    "description": "Verify TransEngCrankEnbl is 1.",
                    "execution": "Read value in receive window.",
                    "actual": "",
                    "expected": "True",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-005",
                    "dcr": "",
                },
                {
                    "step_num": 12,
                    "description": "Open PCAN Explorer and verify PGN ENGSC (F0ED) available and updating.",
                    "execution": "Verify 10F0ED00 is received.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 13,
                    "description": "Verify SPN 7746 byte1 bit position 3(1.3) is 001.",
                    "execution": "EngineStartConsent = consent to operator requested",
                    "actual": "",
                    "expected": "001 - hex\n001 - dec",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-ENG-006",
                    "dcr": "",
                },
                {
                    "step_num": 14,
                    "description": "Stop message and Trace and Save Trace as respective CAN port.",
                    "execution": "Save logs.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
                {
                    "step_num": 15,
                    "description": "Repeat testcase for other applicable CAN port.",
                    "execution": "Repeat steps on next port.",
                    "actual": "",
                    "expected": "",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "",
                    "dcr": "",
                },
            ],
        },
        {
            "case_number": 2,
            "id": "TC_HIL_002",
            "title": "Sample Test",
            "description": "Secondary sample testcase using same template structure.",
            "requirement_id": "REQ-SAMPLE-010",
            "pre_conditions": "1. Device powered.\n2. Communication link established.",
            "post_conditions": "NA",
            "expected_result": "All expected values align with target values.",
            "test_log_path": "",
            "test": "HIL-Open Loop",
            "steps": [
                {
                    "step_num": 1,
                    "description": "Turn on device.",
                    "execution": "Verify power LED turns on.",
                    "actual": "",
                    "expected": "ON",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-SAMPLE-011",
                    "dcr": "",
                },
                {
                    "step_num": 2,
                    "description": "Send frame 0x100.",
                    "execution": "Transmit 100 - hex and observe response.",
                    "actual": "",
                    "expected": "000 - hex",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-SAMPLE-012",
                    "dcr": "",
                },
                {
                    "step_num": 3,
                    "description": "Validate response.",
                    "execution": "Confirm response code equals expected value.",
                    "actual": "",
                    "expected": "PASS",
                    "units": "None",
                    "tolerance": "0",
                    "pass_fail": "PASS/FAIL",
                    "req_id": "REQ-SAMPLE-013",
                    "dcr": "",
                },
            ],
        },
    ]


def apply_cell_style(cell, font, alignment, border, fill=None):
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill is not None:
        cell.fill = fill


def write_test_case(ws, start_row, test_case, styles):
    thin_border = styles["thin_border"]
    label_font = styles["label_font"]
    value_font = styles["value_font"]
    title_font = styles["title_font"]
    left_align = styles["left_align"]
    left_top_wrap = styles["left_top_wrap"]
    center_wrap = styles["center_wrap"]

    ws.cell(row=start_row, column=1, value=f"Test Case #{test_case['case_number']}")
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = title_font
    title_cell.alignment = left_align

    header_fields = [
        ("ID", test_case.get("id", "")),
        ("Title", test_case.get("title", "")),
        ("Description", test_case.get("description", "")),
        ("Requirement ID", test_case.get("requirement_id", "")),
        ("Pre-Conditions", test_case.get("pre_conditions", "")),
        ("Post Conditions", test_case.get("post_conditions", "NA")),
        ("Expected Result", test_case.get("expected_result", "")),
        ("Test Log path", test_case.get("test_log_path", "")),
        ("Test", test_case.get("test", "HIL-Open Loop")),
    ]

    row = start_row + 1
    for label, value in header_fields:
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=value)

        apply_cell_style(label_cell, label_font, left_align, thin_border)
        apply_cell_style(value_cell, value_font, left_top_wrap, thin_border)

        if label == "Pre-Conditions":
            ws.row_dimensions[row].height = 60

        row += 1

    table_header_row = row + 1

    headers = [
        "Test Step #",
        "Test Step Description",
        "Test Execution Steps",
        "Actual Value",
        "Expected Value (Target Value)",
        "Units",
        "Tolerance",
        "Pass/Fail",
        "Req ID [Optional]",
        "DCR for Failed Test Case/Test Step [Optional]",
    ]

    header_fills = {
        1: styles["gray_fill"],
        2: styles["gray_fill"],
        3: styles["gray_fill"],
        4: styles["orange_fill"],
        5: styles["gray_fill"],
        6: styles["gray_fill"],
        7: styles["pink_fill"],
        8: styles["light_orange_fill"],
        9: styles["gray_fill"],
        10: styles["gray_fill"],
    }

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=col_idx, value=header)
        apply_cell_style(cell, label_font, center_wrap, thin_border, fill=header_fills[col_idx])

    data_row = table_header_row + 1
    for step in test_case.get("steps", []):
        values = [
            step.get("step_num", ""),
            step.get("description", ""),
            step.get("execution", ""),
            step.get("actual", ""),
            step.get("expected", ""),
            step.get("units", "None"),
            step.get("tolerance", "0"),
            step.get("pass_fail", "PASS/FAIL"),
            step.get("req_id", ""),
            step.get("dcr", ""),
        ]

        max_lines = 1
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            align = left_top_wrap if col_idx in (2, 3, 4, 5, 9, 10) else left_align
            apply_cell_style(cell, value_font, align, thin_border)

            line_count = str(value).count("\n") + 1 if value not in (None, "") else 1
            max_lines = max(max_lines, line_count)

        ws.row_dimensions[data_row].height = max(22, max_lines * 14)
        data_row += 1

    return data_row + 2


def create_test_plan_workbook(test_cases):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Plan"

    ws.sheet_view.showGridLines = True

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 30

    thin = Side(style="thin", color="000000")

    styles = {
        "thin_border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "title_font": Font(name="Arial", size=12, bold=True, color="000000"),
        "label_font": Font(name="Arial", size=10, bold=True, color="000000"),
        "value_font": Font(name="Arial", size=10, bold=False, color="000000"),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=False),
        "left_top_wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "gray_fill": PatternFill(fill_type="solid", fgColor="D3D3D3"),
        "orange_fill": PatternFill(fill_type="solid", fgColor="FF9900"),
        "pink_fill": PatternFill(fill_type="solid", fgColor="FFB6C1"),
        "light_orange_fill": PatternFill(fill_type="solid", fgColor="FFCC99"),
    }

    current_row = 3
    for test_case in test_cases:
        current_row = write_test_case(ws, current_row, test_case, styles)

    wb.save("test_cases.xlsx")


if __name__ == "__main__":
    test_cases_data = get_dummy_test_cases()
    create_test_plan_workbook(test_cases_data)
    print("Excel file generated successfully.")
